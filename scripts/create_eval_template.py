"""
Create Excel template for human evaluation
Run: python scripts/create_eval_template.py
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

SAMPLES_PATH = "results/human_eval_samples.csv"
TEMPLATE_PATH = "evaluation/human_eval_template.xlsx"

def create_excel_template():
    """Create formatted Excel file for human evaluation"""
    
    # Load samples
    samples = pd.read_csv(SAMPLES_PATH)

    # Create workbook
    wb = Workbook()

    # ================================
    # 1) Instructions Sheet
    # ================================
    ws_instructions = wb.active
    ws_instructions.title = "Instructions"

    instructions = [
        ["Human Evaluation Instructions"],
        [""],
        ["Please rate each model-generated comment based on the following criteria:"],
        [""],
        ["1. Correctness (1–5):"],
        ["  1 = Completely incorrect/broken"],
        ["  2 = Mostly incorrect with major errors"],
        ["  3 = Partially correct, some issues"],
        ["  4 = Mostly correct, minor issues"],
        ["  5 = Completely correct"],
        [""],
        ["2. Readability (1–5):"],
        ["  1 = Unreadable, poorly formatted"],
        ["  2 = Hard to read"],
        ["  3 = Acceptable readability"],
        ["  4 = Good readability"],
        ["  5 = Excellent formatting"],
        [""],
        ["3. Usefulness (1–5):"],
        ["  1 = Not useful at all"],
        ["  2 = Slightly useful"],
        ["  3 = Moderately useful"],
        ["  4 = Very useful"],
        ["  5 = Production-ready"],
        [""],
        ["4. Hallucination Detection:"],
        ["  - Mark Yes/No if the model makes up incorrect facts"],
        ["  - Severity: Minor / Moderate / Critical"],
        ["  - Describe the hallucination briefly"],
    ]

    for row in instructions:
        ws_instructions.append(row)

    ws_instructions["A1"].font = Font(size=14, bold=True)


    # ================================
    # 2) Evaluation Sheet
    # ================================
    ws_eval = wb.create_sheet("Evaluation")

    # Rename columns to what evaluators expect
    samples.rename(columns={
        "input": "code_snippet",
        "expected": "ground_truth",
        "model_output": "generated_comment"
    }, inplace=True)

    # If no ground_truth exists, fill blank
    if "ground_truth" not in samples.columns:
        samples["ground_truth"] = ""

    # Build final evaluation DF
    eval_data = samples[["id", "code_snippet", "generated_comment", "ground_truth"]].copy()

    # Add evaluation columns
    eval_data["Correctness (1–5)"] = ""
    eval_data["Readability (1–5)"] = ""
    eval_data["Usefulness (1–5)"] = ""
    eval_data["Has Hallucination? (Yes/No)"] = ""
    eval_data["Hallucination Severity (Minor/Moderate/Critical)"] = ""
    eval_data["Hallucination Description"] = ""
    eval_data["Notes"] = ""

    # Write table to sheet
    for r_idx, row in enumerate(dataframe_to_rows(eval_data, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_eval.cell(row=r_idx, column=c_idx, value=value)

            # Style header
            if r_idx == 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCE5FF", fill_type="solid")
            
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Set column widths
    widths = {
        "A": 8,   # id
        "B": 50,  # code_snippet
        "C": 50,  # generated_comment
        "D": 50,  # ground_truth
        "E": 18,  # correctness
        "F": 18,  # readability
        "G": 18,  # usefulness
        "H": 18,  # hallucination yes/no
        "I": 25,  # severity
        "J": 40,  # description
        "K": 30,  # notes
    }

    for col, width in widths.items():
        ws_eval.column_dimensions[col].width = width

    # Save file
    wb.save(TEMPLATE_PATH)
    print(f"✓ Created Excel template at: {TEMPLATE_PATH}")
    print(f"✓ Includes {len(samples)} samples")


if __name__ == "__main__":
    import os
    os.makedirs("evaluation", exist_ok=True)
    create_excel_template()
