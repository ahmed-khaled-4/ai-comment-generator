import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

SAMPLES_PATH = "results/human_eval_samples.csv"
TEMPLATE_PATH = "evaluation/human_eval_template.xlsx"

def create_excel_template():
    """Create formatted Excel file for human evaluation"""
    
    # Load samples
    samples = pd.read_csv(SAMPLES_PATH)
    
    # Create workbook
    wb = Workbook()
    
    # Instructions sheet
    ws_instructions = wb.active
    ws_instructions.title = "Instructions"
    
    instructions = [
        ["Human Evaluation Instructions", ""],
        ["", ""],
        ["Please rate each code sample on the following criteria:", ""],
        ["", ""],
        ["1. Correctness (1-5):", ""],
        ["  1 = Completely incorrect/broken", ""],
        ["  2 = Mostly incorrect with major errors", ""],
        ["  3 = Partially correct, some issues", ""],
        ["  4 = Mostly correct, minor issues", ""],
        ["  5 = Completely correct", ""],
        ["", ""],
        ["2. Readability (1-5):", ""],
        ["  1 = Unreadable, poorly formatted", ""],
        ["  2 = Hard to read", ""],
        ["  3 = Acceptable readability", ""],
        ["  4 = Good readability", ""],
        ["  5 = Excellent, professional formatting", ""],
        ["", ""],
        ["3. Usefulness (1-5):", ""],
        ["  1 = Not useful at all", ""],
        ["  2 = Slightly useful", ""],
        ["  3 = Moderately useful", ""],
        ["  4 = Very useful", ""],
        ["  5 = Production-ready", ""],
        ["", ""],
        ["4. Hallucination Detection:", ""],
        ["  - Select if the output contains fabricated/incorrect information", ""],
        ["  - Severity: Minor / Moderate / Critical", ""],
        ["  - Describe the hallucination briefly", ""],
    ]
    
    for row in instructions:
        ws_instructions.append(row)
    
    # Format instructions
    ws_instructions['A1'].font = Font(bold=True, size=14)
    
    # Evaluation sheet
    ws_eval = wb.create_sheet("Evaluation")
    
    # Prepare data
    eval_data = samples[['id', 'input', 'expected', 'model_output', 'task']].copy()
    eval_data['Correctness (1-5)'] = ''
    eval_data['Readability (1-5)'] = ''
    eval_data['Usefulness (1-5)'] = ''
    eval_data['Has Hallucination? (Yes/No)'] = ''
    eval_data['Hallucination Severity (Minor/Moderate/Critical)'] = ''
    eval_data['Hallucination Description'] = ''
    eval_data['Notes'] = ''
    
    # Write to sheet
    for r_idx, row in enumerate(dataframe_to_rows(eval_data, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_eval.cell(row=r_idx, column=c_idx, value=value)
            
            # Format header
            if r_idx == 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
                cell.alignment = Alignment(wrap_text=True, vertical='center')
    
    # Adjust column widths
    column_widths = {
        'A': 8,   # ID
        'B': 50,  # Input
        'C': 50,  # Expected
        'D': 50,  # Model Output
        'E': 15,  # Task
        'F': 15,  # Correctness
        'G': 15,  # Readability
        'H': 15,  # Usefulness
        'I': 20,  # Has Hallucination
        'J': 20,  # Severity
        'K': 40,  # Description
        'L': 30,  # Notes
    }
    
    for col, width in column_widths.items():
        ws_eval.column_dimensions[col].width = width
    
    # Save
    wb.save(TEMPLATE_PATH)
    print(f"✓ Created Excel template: {TEMPLATE_PATH}")
    print(f"  Samples: {len(samples)}")
    print(f"\nShare this file with evaluators:")
    print(f"  {TEMPLATE_PATH}")

if __name__ == "__main__":
    import os
    os.makedirs('evaluation', exist_ok=True)
    create_excel_template()